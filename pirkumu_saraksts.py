import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
Pirkumu_saraksts = {}
def pievienot_pirkuma_sarakstam():

  Produkta_nosaukums = input("Produkta nosaukums: ")
  produkta_kategorija = input("Produkta kategorija: ")
  produkta_cena = float(input("Cena(EUR): "))
  Pirkumu_saraksts[Produkta_nosaukums]={"Produkta kategorija": produkta_kategorija,
                                            "Produkta cena(EUR)": produkta_cena,
                                            "Pirkuma statuss": "Vajag nopirkt"}            
  print(Produkta_nosaukums, "ir pievienots pirkuma sarakstam.")

def dzest_produktu_no_saraksta():
  Produkta_nosaukums = input("Ievadiet produkta nosaukumu, kuru vajag izdzēst: ")
  if Produkta_nosaukums in Pirkumu_saraksts:
    del Pirkumu_saraksts[Produkta_nosaukums]
    print(Produkta_nosaukums, "ir izdzēsts no pirkuma saraksta\n")
  else:
    print("Produkts nav atrasts pirkumu sarakstā.")

def apskatit_sarakstu():
  if not Pirkumu_saraksts:
    print("Pirkumu saraksts ir tukšs\n")
    return
  for Produkta_nosaukums, i in Pirkumu_saraksts.items():
    print(Produkta_nosaukums + ' | ' + i['Produkta kategorija'] + ' | ' + str(i['Produkta cena(EUR)']) + " EUR | " + i['Pirkuma statuss'])
  print()

def perkamo_un_nopirkto_produktu_izmaksas():
  summa_nopirktajiem_produktiem = sum(i['Produkta cena(EUR)'] for i in Pirkumu_saraksts.values() if i["Pirkuma statuss"] == "Nopirkts")
  summa_pērkamajiem_produktiem = sum(i['Produkta cena(EUR)'] for i in Pirkumu_saraksts.values() if i["Pirkuma statuss"] == "Vajag nopirkt")
  print(f"Produkti nopirkti par:{summa_nopirktajiem_produktiem:.2f}EUR")
  print(f"Produkti jāpērk par:{summa_pērkamajiem_produktiem:.2f}EUR\n")

def apstiprinat_ka_nopirkts():
  Produkta_nosaukums = input("Produkta nosaukums, kas ir nopirkts: ")
  if Produkta_nosaukums in Pirkumu_saraksts:
    Pirkumu_saraksts[Produkta_nosaukums]["Pirkuma statuss"]="Nopirkts"
    print(Produkta_nosaukums, "ir nopirkts.\n")
  else:
    print('Produkts nav atrasts pirkumu sarakstā.')

def ierakstit_excel():
  if not Pirkumu_saraksts:
    print("Pirkumu saraksts ir tukšs.\n") 
    return
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Pirkumu saraksts"
  ws.append(["Produkta nosaukums", 'Produkta kategorija', 'Produkta cena(EUR)', "Pirkuma statuss"])
  for Produkta_nosaukums, i in Pirkumu_saraksts.items():
    ws.append([Produkta_nosaukums, i['Produkta kategorija'], i['Produkta cena(EUR)'],i["Pirkuma statuss"]])
  ws.append([])
  ws.append(["Dati ievadīti:",datetime.now().strftime("%H:%M:%S %d-%m-%Y")])
  for kolonna in ws.columns:
    garakā_teksta_virkne = 0
    kolonnas_burts = get_column_letter(kolonna[0].column)
    for šūna in kolonna:
        try:
            if šūna.value:
                garakā_teksta_virkne = max(garakā_teksta_virkne, len(str(šūna.value)))
        except:
            pass
    jaunais_šūnas_platums = garakā_teksta_virkne + 2
    ws.column_dimensions[kolonnas_burts].width = jaunais_šūnas_platums 
  f = "Pirkumu_saraksts.xlsx"
  wb.save(f)
  print("Pirkumu saraksts ir ierakstīts Excel failā ar nosaukumu: ",f)
  print()



def izvelne():
  while True:
    print("*************** Pirkumu saraksts *****************")
    print("1. Pievienot produktu pirkumu sarakstam")
    print("2. Dzēst produktu no pirkumu saraksta")
    print("3. Apskatīt pirkumu sarakstu")
    print("4. Apstriprināt produktu, kā nopirktu")
    print("5. Apskatīt izmaksas nopirktajām un pērkamajām precēm")
    print("6. Ierakstīt Excel failā pirkumu sarakstu")
    print("7. Iziet no programmas")
    print("**************************************************")
    print()
    lietotaja_opcija = input("Izvēlies Opciju no 1 līdz 7: ")

    if lietotaja_opcija == "1":
      pievienot_pirkuma_sarakstam()
    elif lietotaja_opcija == "2":
      dzest_produktu_no_saraksta()
    elif lietotaja_opcija == "3":
      apskatit_sarakstu()
    elif lietotaja_opcija == "4":
      apstiprinat_ka_nopirkts()
    elif lietotaja_opcija == "5":
      perkamo_un_nopirkto_produktu_izmaksas()
    elif lietotaja_opcija == "6":
      ierakstit_excel()
    elif lietotaja_opcija == "7":
      break
    else: print("Ķļūdaina ievade. Jāizvēlas opcijas no 1 līdz 7\n")

if __name__ == "__main__":
  izvelne()


